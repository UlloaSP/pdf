from pathlib import Path

def run(inputs: list[str], output_dir: str, options: dict) -> list[str]:
    if len(inputs) != 1:
        raise ValueError("Selecciona exactamente un archivo.")
    source = Path(inputs[0])
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    import re
    from pypdf import PdfReader
    from openpyxl import Workbook
    pdf = PdfReader(source)
    texts = [(page.extract_text(extraction_mode="layout") or "") if "/Contents" in page else "" for page in pdf.pages]
    if not any(text.strip() for text in texts):
        raise ValueError("No hay texto extraíble. Ejecuta OCR antes de convertir.")
    workbook = Workbook()
    workbook.remove(workbook.active)
    for i, text in enumerate(texts):
        sheet = workbook.create_sheet(f"Página {i + 1}")
        for row, line in enumerate(text.splitlines(), 1):
            for col, value in enumerate(re.split(r"\t+| {2,}", line.strip()), 1):
                cell = sheet.cell(row, col, value)
                cell.data_type = "s"
        sheet.column_dimensions["A"].width = 60
    target = out / "converted.xlsx"
    workbook.save(target)
    return [str(target)]
